from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

refinedData = Workbook()
refinedSheet = refinedData.active

fightersData = load_workbook('Fighters.xlsx')
fightersSheet = fightersData.active

fightsData = load_workbook('Fights.xlsx')
fightsSheet = fightsData.active

# getFighterAge(fid)
# -- given fighter's fid and a date, compute age at that date
def getFighterAge(fid, fightDate):
	age = None

	if fid == None:
		return None

	for i in xrange(2, 1563):
		fighterDOB = None

		if fid == fightersSheet['B' + str(i)].value:
			fighterDOB = fightersSheet['E' + str(i)].value
			break 
	
	if isinstance(fighterDOB, datetime.date):
		return fightDate.year - fighterDOB.year
	elif fighterDOB == None:
		return None
	else:
		return fightDate.year - int(fighterDOB[-4:])

# populateAgeDifference()
# -- populate refined data with age differences

def populateAgeDifference():
	print "Populating age difference..."
	for i in xrange(2, 3571):
		fighterOneAge = getFighterAge(fightsSheet['N' + str(i)].value, fightsSheet['F' + str(i)].value)
		fighterTwoAge = getFighterAge(fightsSheet['O' + str(i)].value, fightsSheet['F' + str(i)].value)
		if fighterOneAge==None or fighterTwoAge==None:
			refinedSheet['A' + str(i)].value = ""
		elif fighterOneAge > fighterTwoAge:
			refinedSheet['A' + str(i)].value = fighterOneAge - fighterTwoAge
		else:
			refinedSheet['A' + str(i)].value = fighterTwoAge - fighterOneAge

# getWeight()
# -- get weight of fighter given fid

def getWeight(fid):
	weight = None
	if fid == None:
		return None
	for i in xrange(2, 1563):
		if fid == fightersSheet['B' + str(i)].value:
			weight = fightersSheet['G' + str(i)].value
			break 
	if weight == None:
		return None
	else:
		return int(weight)


# populateWeightDifference()
# -- populate refined data with weight difference

def populateWeightDifference():
	print "Populating weight difference..."
	for i in xrange(2, 3571):
		fighterOneWeight = getWeight(fightsSheet['N' + str(i)].value)
		fighterTwoWeight = getWeight(fightsSheet['O' + str(i)].value)
		if fighterOneWeight==None or fighterTwoWeight==None:
			refinedSheet['B' + str(i)].value = ""
		elif fighterOneWeight > fighterTwoWeight:
			refinedSheet['B' + str(i)].value = fighterOneWeight - fighterTwoWeight
		else:
			refinedSheet['B' + str(i)].value = fighterTwoWeight - fighterOneWeight

# getHeight()
# -- get height of fighter given fid

def getHeight(fid):
	height = None
	if fid == None:
		return None
	for i in xrange(2, 1563):
		if fid == fightersSheet['B' + str(i)].value:
			height = fightersSheet['F' + str(i)].value
			break 
	if height == None:
		return None
	else:
		return int(height)

# populateHeightDifference()
# -- populate refined data with Height difference

def populateHeightDifference():
	print "Populating height difference..."
	for i in xrange(2, 3571):
		fighterOneHeight = getHeight(fightsSheet['N' + str(i)].value)
		fighterTwoHeight = getHeight(fightsSheet['O' + str(i)].value)
		if fighterOneHeight==None or fighterTwoHeight==None:
			refinedSheet['C' + str(i)].value = ""
		elif fighterOneHeight > fighterTwoHeight:
			refinedSheet['C' + str(i)].value = fighterOneHeight - fighterTwoHeight
		else:
			refinedSheet['C' + str(i)].value = fighterTwoHeight - fighterOneHeight


# populateRefs()
#

def populateRefs():
	print "Populating refs..."
	for i in xrange(2, 3571):
		refinedSheet['D' + str(i)].value = fightsSheet['R' + str(i)].value

# populateRound()
#

def populateRound():
	print "Populating round..."
	for i in xrange(2, 3571):
		refinedSheet['E' + str(i)].value = fightsSheet['S' + str(i)].value

# getClass()
# -- gets class from fid

def getClass(fid):
	fighterClass = None
	if fid == None:
		return None
	for i in xrange(2, 1563):
		if fid == fightersSheet['B' + str(i)].value:
			fighterClass = fightersSheet['I' + str(i)].value
			break 
	if fighterClass == None:
		return None
	else:
		return fighterClass

# populateClass()
#

def populateClass():
	print "Populating class..."
	for i in xrange(2, 3571):
		fighterClass = getClass(fightsSheet['N' + str(i)].value)
		if fighterClass==None:
			refinedSheet['F' + str(i)].value = ""
		else:
			refinedSheet['F' + str(i)].value = fighterClass

# populateMethod()
#

def populateMethod():
	print "Populating method..."
	for i in xrange(2, 3571):
		refinedSheet['G' + str(i)].value = fightsSheet['P' + str(i)].value


def main():
	populateAgeDifference()
	populateWeightDifference()
	populateHeightDifference()
	populateRefs()
	populateRound()
	populateClass()
	populateMethod()
	refinedData.save('refinedData.xlsx')

main()